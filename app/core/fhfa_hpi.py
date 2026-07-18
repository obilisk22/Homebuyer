"""FHFA ZIP5 annual HPI download + decade CAGR."""

from __future__ import annotations

import time
from pathlib import Path

import requests
from openpyxl import load_workbook

from app.core.overlay_cache import cache_dir

FHFA_ZIP5_URL = "https://www.fhfa.gov/hpi/download/annual/hpi_at_zip5.xlsx"
CACHE_TTL_S = 30 * 24 * 3600
REQUEST_TIMEOUT_S = 60
_NAMESPACE = "fhfa"


def cagr_from_index_series(
    points: list[tuple[int, float]], *, span_years: int = 10
) -> float | None:
    if not points or span_years <= 0:
        return None
    by_year = {int(y): float(idx) for y, idx in points if idx and float(idx) > 0}
    if not by_year:
        return None
    end_year = max(by_year)
    start_year = end_year - span_years
    if start_year not in by_year or end_year not in by_year:
        # nearest available start at or before target year if exact miss
        candidates = [y for y in by_year if y <= start_year]
        if not candidates:
            return None
        start_year = max(candidates)
        actual_span = end_year - start_year
        if actual_span <= 0:
            return None
    else:
        actual_span = span_years
    start_idx = by_year[start_year]
    end_idx = by_year[end_year]
    if start_idx <= 0 or end_idx <= 0:
        return None
    return ((end_idx / start_idx) ** (1.0 / actual_span) - 1.0) * 100.0


def _cached_xlsx_path() -> Path:
    return cache_dir(_NAMESPACE) / "hpi_at_zip5.xlsx"


def ensure_zip5_workbook() -> Path | None:
    path = _cached_xlsx_path()
    if path.is_file():
        age = time.time() - path.stat().st_mtime
        if age <= CACHE_TTL_S:
            return path
    try:
        resp = requests.get(FHFA_ZIP5_URL, timeout=REQUEST_TIMEOUT_S)
        resp.raise_for_status()
        path.write_bytes(resp.content)
        return path
    except (OSError, requests.RequestException):
        return path if path.is_file() else None


_index_cache: dict[str, list[tuple[int, float]]] | None = None
_index_cache_mtime: float | None = None


def _header_columns(header: tuple[object, ...] | list[object]) -> tuple[int, int, int] | None:
    cols = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
    zip_i = next((cols[name] for name in cols if "zip" in name), None)
    year_i = next((cols[name] for name in cols if "year" in name), None)
    hpi_names = [
        name
        for name in cols
        if "hpi" in name and "annual change" not in name and "%" not in name
    ]
    idx_i = cols.get("hpi") if "hpi" in cols else (cols[hpi_names[0]] if hpi_names else None)
    if zip_i is None or year_i is None or idx_i is None:
        return None
    return zip_i, year_i, idx_i


def _index_from_worksheet(ws) -> dict[str, list[tuple[int, float]]]:
    """Parse all ZIP series after locating the FHFA header (skips title rows)."""
    rows = ws.iter_rows(values_only=True)
    zip_i = year_i = idx_i = None
    for _ in range(25):
        header = next(rows, None)
        if not header:
            continue
        found = _header_columns(header)
        if found is None:
            continue
        zip_i, year_i, idx_i = found
        break
    if zip_i is None or year_i is None or idx_i is None:
        return {}

    out: dict[str, list[tuple[int, float]]] = {}
    required_length = max(zip_i, year_i, idx_i) + 1
    for row in rows:
        if not row or len(row) < required_length or row[zip_i] is None:
            continue
        z = "".join(c for c in str(row[zip_i]) if c.isdigit()).zfill(5)
        if len(z) < 5:
            continue
        z = z[:5]
        try:
            year = int(row[year_i])
            idx = float(row[idx_i])
        except (TypeError, ValueError):
            continue
        if idx <= 0:
            continue
        out.setdefault(z, []).append((year, idx))
    return out


def _series_from_worksheet(ws, zip5: str) -> list[tuple[int, float]]:
    """Extract one ZIP's annual HPI levels (used by tests with small workbooks)."""
    return list(_index_from_worksheet(ws).get(zip5, []))


def _load_zip_index() -> dict[str, list[tuple[int, float]]]:
    """Load and memoize the full ZIP→HPI index for the cached workbook."""
    global _index_cache, _index_cache_mtime
    path = ensure_zip5_workbook()
    if path is None:
        return {}
    mtime = path.stat().st_mtime
    if _index_cache is not None and _index_cache_mtime == mtime:
        return _index_cache
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        _index_cache = _index_from_worksheet(wb.active)
        _index_cache_mtime = mtime
        return _index_cache
    finally:
        wb.close()


def _load_zip_series(zip_code: str) -> list[tuple[int, float]]:
    zip5 = "".join(c for c in str(zip_code) if c.isdigit())[:5]
    if len(zip5) < 5:
        return []
    return list(_load_zip_index().get(zip5.zfill(5), []))


def zip5_cagr(zip_code: str, *, span_years: int = 10) -> float | None:
    try:
        return cagr_from_index_series(_load_zip_series(zip_code), span_years=span_years)
    except Exception:
        return None
